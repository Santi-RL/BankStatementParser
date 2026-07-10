from collections import Counter
from numbers import Real
from typing import Any, Callable, Dict, List

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reconciliation import aggregate_reconciliation_status


RECONCILIATION_COLUMNS = [
    ("source_file", "Archivo"),
    ("bank", "Banco"),
    ("scope_label", "Cuenta o producto"),
    ("currency", "Moneda"),
    ("period_start", "Período desde"),
    ("period_end", "Período hasta"),
    ("opening_balance", "Saldo inicial"),
    ("credits", "Créditos informados"),
    ("debits", "Débitos informados"),
    ("net_movements", "Neto de movimientos"),
    ("calculated_closing_balance", "Saldo final calculado"),
    ("closing_balance", "Saldo final informado"),
    ("difference", "Diferencia"),
    ("status", "Estado"),
    ("reason", "Observación"),
]

STATUS_LABELS = {
    "passed": "Conciliado",
    "failed": "Con diferencia",
    "not_available": "No disponible",
    "partial": "Parcial",
}

REASON_LABELS = {
    "balanced": "Sin diferencias.",
    "difference": "El saldo final informado no coincide con el saldo calculado.",
    "not_supported": "Conciliación no disponible para este formato.",
    "summary_not_found": "No se encontraron saldos inicial y final utilizables.",
    "missing_summary_values": "Faltan valores del resumen necesarios para conciliar.",
    "invalid_status": "El resultado de conciliación no es válido.",
}


def _display_value(key: str, value: Any) -> Any:
    if key == "status":
        return STATUS_LABELS.get(str(value), "No disponible")
    if key == "reason":
        return REASON_LABELS.get(str(value), str(value or ""))
    return value


def create_reconciliation_sheet(
    workbook,
    reconciliations: List[Dict[str, Any]],
    safe_value: Callable[[Any], Any],
) -> None:
    worksheet = workbook.create_sheet("Conciliación", 1)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for column, (_, label) in enumerate(RECONCILIATION_COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=column, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    rows = reconciliations or [
        {
            "status": "not_available",
            "reason": "not_supported",
        }
    ]
    numeric_keys = {
        "opening_balance",
        "credits",
        "debits",
        "net_movements",
        "calculated_closing_balance",
        "closing_balance",
        "difference",
    }
    status_column = next(
        index
        for index, (key, _) in enumerate(RECONCILIATION_COLUMNS, start=1)
        if key == "status"
    )

    for row_index, reconciliation in enumerate(rows, start=2):
        for column, (key, _) in enumerate(RECONCILIATION_COLUMNS, start=1):
            value = safe_value(_display_value(key, reconciliation.get(key, "")))
            cell = worksheet.cell(row=row_index, column=column, value=value)
            if key in numeric_keys and isinstance(value, Real) and not isinstance(value, bool):
                cell.number_format = "#,##0.00"

        status = reconciliation.get("status", "not_available")
        status_cell = worksheet.cell(row=row_index, column=status_column)
        if status == "passed":
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif status == "failed":
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for column in worksheet.columns:
        column_letter = get_column_letter(column[0].column)
        max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 48)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def append_reconciliation_summary(
    worksheet,
    reconciliations: List[Dict[str, Any]],
    start_row: int,
    safe_value: Callable[[Any], Any],
) -> int:
    counts = Counter(
        reconciliation.get("status", "not_available")
        for reconciliation in reconciliations
    )
    overall_status = aggregate_reconciliation_status(reconciliations)

    worksheet[f"A{start_row}"] = "Control de conciliación"
    worksheet[f"A{start_row}"].font = Font(bold=True, size=12)
    rows = [
        ("Estado general:", STATUS_LABELS.get(overall_status, "No disponible")),
        ("Conciliadas:", counts.get("passed", 0)),
        ("Con diferencias:", counts.get("failed", 0)),
        ("No disponibles:", counts.get("not_available", 0)),
    ]
    for offset, (label, value) in enumerate(rows, start=1):
        worksheet[f"A{start_row + offset}"] = label
        worksheet[f"B{start_row + offset}"] = safe_value(value)
    return start_row + len(rows)
