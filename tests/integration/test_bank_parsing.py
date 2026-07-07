import io
import re
from pathlib import Path

import pytest
from openpyxl import load_workbook

from excel_generator import ExcelGenerator
from pdf_processor import PDFProcessor




class FixtureTextProcessor(PDFProcessor):
    def __init__(self, text: str):
        super().__init__()
        self._text = text

    def _extract_text_from_pdf(self, file_path, debug_log=None):
        if debug_log is not None:
            debug_log.append("fixture extractor")
        return self._text


MULTI_SCOPE_FIXTURE_CASES = [
    pytest.param(
        {
            "bank_id": "bbva",
            "format_id": "default",
            "fixture_path": Path("parser_specs/bbva/default/fixtures/sample_text.txt"),
            "file_name": "01-2023 BBVA sanitized fixture.txt",
            "scope_count": 5,
            "expected_scope_ids": {
                "bank_account:CA_$_354-428727/2",
                "bank_account:CA_U$S_354-412259/7",
                "bank_account:CC_$_354-560404/1",
                "credit_card:0837946968",
                "debit_card:0429",
            },
            "expected_scope_currencies": {"ARS", "USD"},
            "all_transaction_count": 6,
            "all_transaction_labels": {
                "CA $ 354-428727/2",
                "CC $ 354-560404/1",
                "Tarjeta crédito 0837946968",
                "Visa Débito 0429",
            },
            "group_selection_scope_ids": [
                "bank_account:CA_$_354-428727/2",
                "bank_account:CA_U$S_354-412259/7",
                "bank_account:CC_$_354-560404/1",
            ],
            "group_selection_product_type": "bank_account",
            "group_selection_labels": {"CA $ 354-428727/2", "CC $ 354-560404/1"},
            "group_selection_count": 2,
            "single_scope_id": "credit_card:0837946968",
            "single_scope_product_type": "credit_card",
            "single_scope_count": 3,
        },
        id="bbva_consolidated",
    ),
    pytest.param(
        {
            "bank_id": "brubank",
            "format_id": "default",
            "fixture_path": Path("parser_specs/brubank/default/fixtures/multi_account_sample_text.txt"),
            "file_name": "brubank-multi-account sanitized fixture.txt",
            "scope_count": 3,
            "expected_scope_ids": {
                "bank_account:caja_ahorro_ars",
                "bank_account:caja_ahorro_usd",
                "bank_account:cuenta_remunerada_ars",
            },
            "expected_scope_currencies": {"ARS", "USD"},
            "all_transaction_count": 12,
            "all_transaction_labels": {"Caja de ahorro ARS", "Cuenta remunerada ARS"},
            "group_selection_scope_ids": [
                "bank_account:caja_ahorro_ars",
                "bank_account:cuenta_remunerada_ars",
            ],
            "group_selection_product_type": "bank_account",
            "group_selection_labels": {"Caja de ahorro ARS", "Cuenta remunerada ARS"},
            "group_selection_count": 12,
            "single_scope_id": "bank_account:cuenta_remunerada_ars",
            "single_scope_product_type": "bank_account",
            "single_scope_count": 6,
        },
        id="brubank_multi_account",
    ),
]


def _processor_for_multi_scope_case(case: dict) -> FixtureTextProcessor:
    return FixtureTextProcessor(case["fixture_path"].read_text(encoding="utf-8"))


def _analyze_multi_scope_case(case: dict) -> tuple[FixtureTextProcessor, dict]:
    processor = _processor_for_multi_scope_case(case)
    analysis = processor.analyze_pdf("fixture.pdf", case["file_name"])
    return processor, analysis


@pytest.mark.parametrize("case", MULTI_SCOPE_FIXTURE_CASES)
def test_multi_scope_fixture_analysis_reports_scopes(case):
    _, analysis = _analyze_multi_scope_case(case)

    assert analysis["success"] is True
    assert analysis["bank_detected"] == case["bank_id"]
    assert analysis["format_id"] == case["format_id"]
    assert analysis["multi_scope"] is True
    assert len(analysis["available_scopes"]) == case["scope_count"]
    assert {scope["id"] for scope in analysis["available_scopes"]} == case["expected_scope_ids"]
    assert {scope["currency"] for scope in analysis["available_scopes"]} == case["expected_scope_currencies"]


@pytest.mark.parametrize("case", MULTI_SCOPE_FIXTURE_CASES)
def test_multi_scope_fixture_requires_selection_and_processes_all_selected(case):
    processor, analysis = _analyze_multi_scope_case(case)
    selected_scope_ids = [scope["id"] for scope in analysis["available_scopes"]]

    without_selection = processor.process_pdf("fixture.pdf", case["file_name"])
    with_selection = processor.process_pdf(
        "fixture.pdf",
        case["file_name"],
        selected_scope_ids=selected_scope_ids,
    )

    assert without_selection["success"] is False
    assert without_selection["parse_status"] == "validation_failed"
    assert without_selection["multi_scope"] is True

    assert with_selection["success"] is True
    assert with_selection["parse_status"] == "ok"
    assert with_selection["multi_scope"] is True
    assert with_selection["total_transactions"] == case["all_transaction_count"]
    assert {transaction["scope_label"] for transaction in with_selection["transactions"]} == case[
        "all_transaction_labels"
    ]


@pytest.mark.parametrize("case", MULTI_SCOPE_FIXTURE_CASES)
def test_multi_scope_fixture_filters_selected_scope_group(case):
    processor = _processor_for_multi_scope_case(case)

    result = processor.process_pdf(
        "fixture.pdf",
        case["file_name"],
        selected_scope_ids=case["group_selection_scope_ids"],
    )

    assert result["success"] is True
    assert result["total_transactions"] == case["group_selection_count"]
    assert all(
        transaction["product_type"] == case["group_selection_product_type"]
        for transaction in result["transactions"]
    )
    assert {transaction["scope_label"] for transaction in result["transactions"]} == case[
        "group_selection_labels"
    ]


@pytest.mark.parametrize("case", MULTI_SCOPE_FIXTURE_CASES)
def test_multi_scope_fixture_filters_single_scope(case):
    processor = _processor_for_multi_scope_case(case)

    result = processor.process_pdf(
        "fixture.pdf",
        case["file_name"],
        selected_scope_ids=[case["single_scope_id"]],
    )

    assert result["success"] is True
    assert result["total_transactions"] == case["single_scope_count"]
    assert all(transaction["scope_id"] == case["single_scope_id"] for transaction in result["transactions"])
    assert all(
        transaction["product_type"] == case["single_scope_product_type"]
        for transaction in result["transactions"]
    )


@pytest.mark.parametrize("case", MULTI_SCOPE_FIXTURE_CASES)
def test_multi_scope_fixture_excel_generation_succeeds(case):
    processor, analysis = _analyze_multi_scope_case(case)
    selected_scope_ids = [scope["id"] for scope in analysis["available_scopes"]]
    result = processor.process_pdf(
        "fixture.pdf",
        case["file_name"],
        selected_scope_ids=selected_scope_ids,
    )

    summary = {
        "total_files": 1,
        "successful_files": 1,
        "failed_files": 0,
        "total_transactions": result["total_transactions"],
        "banks_detected": {result["bank_name"]},
        "errors": [],
        "selected_scopes": analysis["available_scopes"],
    }
    workbook_bytes = ExcelGenerator().create_excel_file(result["transactions"], summary)
    workbook = load_workbook(io.BytesIO(workbook_bytes))
    scope_sheet_names = [
        name
        for name in workbook.sheetnames
        if name not in {"Resumen", "Movimientos", "Análisis", "Resumen Mensual"}
    ]
    transaction_scope_groups = {
        (transaction["bank"], transaction["scope_label"])
        for transaction in result["transactions"]
    }

    assert result["success"] is True
    assert "Resumen" in workbook.sheetnames
    assert "Movimientos" in workbook.sheetnames
    assert len(scope_sheet_names) == len(transaction_scope_groups)
    assert all(re.search(r'[:\\/?*\[\]]', name) is None for name in workbook.sheetnames)

def test_galicia_pdf_uses_declarative_spec():
    processor = PDFProcessor()
    result = processor.process_pdf("attached_assets/TestGalicia.pdf", "TestGalicia.pdf")

    assert result["success"] is True
    assert result["parse_status"] == "ok"
    assert result["format_id"] == "default"
    assert result["bank_detected"] == "galicia_ar"
    assert result["total_transactions"] == 52
    assert "Consolidado de retención" not in result["transactions"][-1]["description"]
    assert "Los depósitos en pesos" not in result["transactions"][-1]["description"]


def test_roela_pdf_uses_declarative_spec_successfully():
    processor = PDFProcessor()
    result = processor.process_pdf("attached_assets/BancoRoela.Argentina.Test.pdf", "BancoRoela.Argentina.Test.pdf")

    assert result["success"] is True
    assert result["parse_status"] == "ok"
    assert result["bank_detected"] == "roela_ar"
    assert result["format_id"] == "default"
    assert result["format_version"] == 1
    assert result["total_transactions"] >= 4900
    assert result["transactions"][0]["account"] == "22537/8"
    assert result["transactions"][5]["amount"] < 0


def test_chase_pdf_uses_declarative_spec_with_correct_dates_and_account():
    processor = PDFProcessor()
    result = processor.process_pdf("attached_assets/BANCO CH 2024 1.pdf", "BANCO CH 2024 1.pdf")

    assert result["success"] is True
    assert result["parse_status"] == "ok"
    assert result["bank_detected"] == "chase"
    assert result["format_id"] == "default"
    assert result["format_version"] == 1
    assert result["total_transactions"] == 12
    assert result["transactions"][0]["date"] == "2024-01-02"
    assert result["transactions"][4]["date"] == "2024-01-03"
    assert result["transactions"][0]["account"] == "000000771927196"


def test_chase_second_pdf_matches_same_published_spec():
    processor = PDFProcessor()
    result = processor.process_pdf("attached_assets/BANCO CH 2024 2.pdf", "BANCO CH 2024 2.pdf")

    assert result["success"] is True
    assert result["parse_status"] == "ok"
    assert result["bank_detected"] == "chase"
    assert result["format_id"] == "default"
    assert result["total_transactions"] == 11
    assert result["transactions"][0]["date"] == "2024-02-05"
    assert result["transactions"][-1]["date"] == "2024-02-29"
    assert result["transactions"][0]["account"] == "000000771927196"


def test_bbva_account_summary_pdf_uses_new_single_scope_spec():
    processor = PDFProcessor()
    analysis = processor.analyze_pdf(
        "attached_assets/nuevo_formato/BBVA/Resumen caja de ahorro BBVA 09-2023.pdf",
        "Resumen caja de ahorro BBVA 09-2023.pdf",
    )
    result = processor.process_pdf(
        "attached_assets/nuevo_formato/BBVA/Resumen caja de ahorro BBVA 09-2023.pdf",
        "Resumen caja de ahorro BBVA 09-2023.pdf",
    )

    assert analysis["success"] is True
    assert analysis["bank_detected"] == "bbva"
    assert analysis["format_id"] == "account_summary"
    assert analysis["multi_scope"] is False
    assert len(analysis["available_scopes"]) == 1
    assert analysis["available_scopes"][0]["label"] == "CA $ 354-428727/2"

    assert result["success"] is True
    assert result["parse_status"] == "ok"
    assert result["bank_detected"] == "bbva"
    assert result["format_id"] == "account_summary"
    assert result["format_version"] == 1
    assert result["total_transactions"] == 12
    assert result["transactions"][0]["date"] == "2023-08-22"
    assert result["transactions"][0]["amount"] == 71093.08
    assert result["transactions"][-1]["description"] == "CUENTA VISA NRO. 79083794696899"
    assert all(transaction["scope_label"] == "CA $ 354-428727/2" for transaction in result["transactions"])


def test_mercado_pago_pdf_uses_new_wallet_spec():
    processor = PDFProcessor()
    analysis = processor.analyze_pdf(
        "attached_assets/nuevo_formato/Mercado Pago/Resumen de cuenta Mercado Pago 02-2023.pdf",
        "Resumen de cuenta Mercado Pago 02-2023.pdf",
    )
    result = processor.process_pdf(
        "attached_assets/nuevo_formato/Mercado Pago/Resumen de cuenta Mercado Pago 02-2023.pdf",
        "Resumen de cuenta Mercado Pago 02-2023.pdf",
    )

    assert analysis["success"] is True
    assert analysis["bank_detected"] == "mercado_pago"
    assert analysis["format_id"] == "default"
    assert analysis["multi_scope"] is False
    assert len(analysis["available_scopes"]) == 1
    assert analysis["available_scopes"][0]["label"] == "CVU 0000003100043499119011"

    assert result["success"] is True
    assert result["parse_status"] == "ok"
    assert result["bank_detected"] == "mercado_pago"
    assert result["format_id"] == "default"
    assert result["format_version"] == 1
    assert result["total_transactions"] == 232
    assert result["transactions"][0]["date"] == "2023-02-01"
    assert result["transactions"][0]["description"] == "Transferencia enviada Totoreño"
    assert result["transactions"][0]["account"] == "0000003100043499119011"
    assert result["transactions"][111]["amount"] == 7830.0
    assert result["transactions"][-1]["description"] == "Rendimientos"
