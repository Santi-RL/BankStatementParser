import io
from pathlib import Path

from openpyxl import load_workbook

from excel_generator import ExcelGenerator
from pdf_processor import PDFProcessor


class FixtureProcessor(PDFProcessor):
    def __init__(self, text: str):
        super().__init__()
        self._text = text

    def _extract_text_from_pdf(self, file_path, debug_log=None):
        return self._text


def _fixture(bank_id: str, name: str = "sample_text.txt") -> str:
    return Path(f"parser_specs/{bank_id}/default/fixtures/{name}").read_text(
        encoding="utf-8"
    )


def _summary(total_transactions: int) -> dict:
    return {
        "total_files": 1,
        "successful_files": 1,
        "failed_files": 0,
        "total_transactions": total_transactions,
        "banks_detected": {"Brubank"},
        "errors": [],
        "selected_scopes": [],
    }


def test_brubank_reconciles_each_selected_scope():
    processor = FixtureProcessor(_fixture("brubank", "multi_account_sample_text.txt"))
    analysis = processor.analyze_pdf("fixture.pdf", "brubank.pdf")
    selected_scope_ids = [scope["id"] for scope in analysis["available_scopes"]]

    result = processor.process_pdf(
        "fixture.pdf",
        "brubank.pdf",
        selected_scope_ids=selected_scope_ids,
    )

    assert result["success"] is True
    assert result["parse_status"] == "ok"
    assert result["reconciliation_status"] == "passed"
    assert len(result["reconciliations"]) == 3
    assert {record["status"] for record in result["reconciliations"]} == {"passed"}
    assert {record["period_start"] for record in result["reconciliations"]} == {"2026-06-01"}
    assert {record["period_end"] for record in result["reconciliations"]} == {"2026-06-30"}
    assert {record["difference"] for record in result["reconciliations"]} == {0.0}


def test_brubank_reconciliation_respects_scope_selection():
    processor = FixtureProcessor(_fixture("brubank", "multi_account_sample_text.txt"))
    scope_id = "bank_account:cuenta_remunerada_ars"

    result = processor.process_pdf(
        "fixture.pdf",
        "brubank.pdf",
        selected_scope_ids=[scope_id],
    )

    assert result["success"] is True
    assert result["reconciliation_status"] == "passed"
    assert [record["scope_id"] for record in result["reconciliations"]] == [scope_id]


def test_missing_summary_for_one_scope_reports_partial_without_blocking():
    text = _fixture("brubank", "multi_account_sample_text.txt").replace(
        "Saldo Final $ 2.420,00",
        "Saldo de cierre $ 2.420,00",
        1,
    )
    processor = FixtureProcessor(text)
    analysis = processor.analyze_pdf("fixture.pdf", "brubank.pdf")

    result = processor.process_pdf(
        "fixture.pdf",
        "brubank.pdf",
        selected_scope_ids=[scope["id"] for scope in analysis["available_scopes"]],
    )

    assert result["success"] is True
    assert result["parse_status"] == "ok"
    assert result["reconciliation_status"] == "partial"
    assert len(result["reconciliations"]) == 3
    assert {record["status"] for record in result["reconciliations"]} == {
        "passed",
        "not_available",
    }


def test_reconciliation_difference_does_not_block_processing():
    text = _fixture("brubank", "multi_account_sample_text.txt").replace(
        "Saldo Final $ 2.420,00",
        "Saldo Final $ 2.421,00",
        1,
    )
    processor = FixtureProcessor(text)
    analysis = processor.analyze_pdf("fixture.pdf", "brubank.pdf")

    result = processor.process_pdf(
        "fixture.pdf",
        "brubank.pdf",
        selected_scope_ids=[scope["id"] for scope in analysis["available_scopes"]],
    )

    assert result["success"] is True
    assert result["parse_status"] == "ok"
    assert result["total_transactions"] == 12
    assert result["reconciliation_status"] == "failed"
    failed = [record for record in result["reconciliations"] if record["status"] == "failed"]
    assert len(failed) == 1
    assert failed[0]["difference"] == 1.0


def test_unsupported_format_reports_not_available_without_blocking():
    result = FixtureProcessor(_fixture("galicia_ar")).process_pdf(
        "fixture.pdf",
        "galicia.pdf",
    )

    assert result["success"] is True
    assert result["parse_status"] == "ok"
    assert result["reconciliation_status"] == "not_available"
    assert result["reconciliations"][0]["status"] == "not_available"
    assert result["reconciliations"][0]["reason"] == "not_supported"


def test_excel_contains_safe_reconciliation_detail_and_summary():
    reconciliation = {
        "source_file": "=extracto.pdf",
        "bank": "Brubank",
        "scope_label": "+Caja de ahorro",
        "currency": "ARS",
        "period_start": "2026-06-01",
        "period_end": "2026-06-30",
        "opening_balance": 10.0,
        "credits": 3210.0,
        "debits": 800.0,
        "net_movements": 2410.0,
        "calculated_closing_balance": 2420.0,
        "closing_balance": 2420.0,
        "difference": 0.0,
        "status": "passed",
        "reason": "balanced",
    }
    transactions = [
        {
            "date": "2026-06-01",
            "description": "Movimiento",
            "amount": 2410.0,
            "transaction_type": "credit",
            "bank": "Brubank",
            "currency": "ARS",
        }
    ]

    workbook_bytes = ExcelGenerator().create_excel_file(
        transactions,
        _summary(1),
        [reconciliation],
    )
    workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=False)

    assert workbook.sheetnames[:3] == ["Resumen", "Conciliación", "Movimientos"]
    worksheet = workbook["Conciliación"]
    headers = [cell.value for cell in worksheet[1]]
    assert worksheet.cell(2, headers.index("Archivo") + 1).value == "'=extracto.pdf"
    assert worksheet.cell(2, headers.index("Cuenta o producto") + 1).value == "'+Caja de ahorro"
    assert worksheet.cell(2, headers.index("Estado") + 1).value == "Conciliado"
    assert worksheet.cell(2, headers.index("Diferencia") + 1).value == 0.0
    assert all(cell.data_type != "f" for row in worksheet.iter_rows() for cell in row)

    summary_values = [cell.value for row in workbook["Resumen"].iter_rows() for cell in row]
    assert "Control de conciliación" in summary_values
    assert "Conciliado" in summary_values
