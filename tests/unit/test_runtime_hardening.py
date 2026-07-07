import csv
import io
from pathlib import Path
import tomllib

import pytest
from openpyxl import load_workbook

import pdf_processor as pdf_processor_module
from app import _build_csv_export, _display_total_amount
from excel_generator import ExcelGenerator
from pdf_processor import PDFProcessor
from utils import escape_spreadsheet_formula_value, temporary_pdf_copy


def _build_summary(total_transactions: int, selected_scopes: list[dict] | None = None) -> dict:
    return {
        "total_files": 1,
        "successful_files": 1,
        "failed_files": 0,
        "total_transactions": total_transactions,
        "banks_detected": {"bbva"},
        "errors": [],
        "selected_scopes": selected_scopes or [],
    }


def _find_summary_value(workbook, label: str):
    worksheet = workbook["Resumen"]
    for row in worksheet.iter_rows(min_col=1, max_col=2):
        if row[0].value == label:
            return row[1].value
    raise AssertionError(f"Label not found in Resumen sheet: {label}")


@pytest.mark.parametrize("value", ["=1+1", "+1", "-1", "@cmd", "\tcmd", "\rcmd", "\ncmd"])
def test_spreadsheet_formula_escape_covers_all_formula_triggers(value):
    assert escape_spreadsheet_formula_value(value) == f"'{value}"


class CountingProcessor(PDFProcessor):
    def __init__(self, text: str):
        super().__init__()
        self._text = text
        self.extract_calls = 0

    def _extract_text_from_pdf(self, file_path, debug_log=None):
        self.extract_calls += 1
        if debug_log is not None:
            debug_log.append("counting extractor")
        return self._text


def test_process_pdf_reuses_single_text_extraction():
    text = Path("parser_specs/galicia_ar/default/fixtures/sample_text.txt").read_text(encoding="utf-8")
    processor = CountingProcessor(text)

    result = processor.process_pdf("dummy.pdf", "dummy.pdf", debug=True)

    assert result["success"] is True
    assert processor.extract_calls == 1
    assert "counting extractor" in result["debug_log"]


def test_extract_text_falls_back_to_pypdf(monkeypatch, tmp_path):
    pdf_path = tmp_path / "fallback.pdf"
    pdf_path.write_bytes(b"%PDF-1.4 fallback")

    class DummyPage:
        def extract_text(self):
            return "Fallback text from pypdf"

    class DummyReader:
        def __init__(self, file_obj):
            self.pages = [DummyPage()]

    processor = PDFProcessor()
    debug_log: list[str] = []

    monkeypatch.setattr(pdf_processor_module, "pdfplumber", None)
    monkeypatch.setattr(pdf_processor_module, "PdfReader", DummyReader)

    extracted_text = processor._extract_text_from_pdf(str(pdf_path), debug_log)

    assert "Fallback text from pypdf" in extracted_text
    assert "pypdf succeeded" in debug_log


def test_temporary_pdf_copy_cleans_up_after_exception():
    temp_path: Path | None = None

    with pytest.raises(RuntimeError):
        with temporary_pdf_copy(b"%PDF-1.4 test payload") as generated_path:
            temp_path = Path(generated_path)
            assert temp_path.exists()
            raise RuntimeError("boom")

    assert temp_path is not None
    assert not temp_path.exists()


def test_display_total_amount_uses_transaction_currency():
    transactions = [
        {"amount": 100.0, "currency": "ARS"},
        {"amount": -40.0, "currency": "ARS"},
    ]

    assert _display_total_amount(transactions) == "AR$60.00"


def test_display_total_amount_flags_multiple_currencies():
    transactions = [
        {"amount": 100.0, "currency": "ARS"},
        {"amount": -40.0, "currency": "USD"},
    ]

    assert _display_total_amount(transactions) == "No aplica (múltiples monedas)"


def test_excel_scope_sheet_titles_are_sanitized_and_deduplicated():
    transactions = [
        {
            "date": "2024-01-01",
            "description": "Ingreso",
            "amount": 10.0,
            "balance": 10.0,
            "transaction_type": "Credit",
            "bank": "bbva",
            "account": "1",
            "currency": "EUR",
            "scope_label": "Cuenta / Uno",
            "product_type": "bank_account",
            "linked_account": "",
            "source_file": "bbva.pdf",
        },
        {
            "date": "2024-01-02",
            "description": "Debito",
            "amount": -5.0,
            "balance": 5.0,
            "transaction_type": "Debit",
            "bank": "bbva",
            "account": "2",
            "currency": "EUR",
            "scope_label": "Cuenta : Uno",
            "product_type": "bank_account",
            "linked_account": "",
            "source_file": "bbva.pdf",
        },
    ]

    workbook_bytes = ExcelGenerator().create_excel_file(transactions, _build_summary(total_transactions=2))
    workbook = load_workbook(io.BytesIO(workbook_bytes))

    assert "bbva Cuenta Uno" in workbook.sheetnames
    assert "bbva Cuenta Uno 2" in workbook.sheetnames
    assert all(all(character not in name for character in r'[]:*?/\\') for name in workbook.sheetnames)


def test_excel_summary_flags_multiple_currencies():
    transactions = [
        {
            "date": "2024-01-01",
            "description": "Ingreso",
            "amount": 10.0,
            "balance": 10.0,
            "transaction_type": "Credit",
            "bank": "bbva",
            "account": "1",
            "currency": "EUR",
        },
        {
            "date": "2024-01-02",
            "description": "Debito",
            "amount": -5.0,
            "balance": 5.0,
            "transaction_type": "Debit",
            "bank": "bbva",
            "account": "1",
            "currency": "USD",
        },
    ]

    workbook_bytes = ExcelGenerator().create_excel_file(transactions, _build_summary(total_transactions=2))
    workbook = load_workbook(io.BytesIO(workbook_bytes))

    assert _find_summary_value(workbook, "Total de créditos:") == "No aplica (múltiples monedas)"
    assert _find_summary_value(workbook, "Total de débitos:") == "No aplica (múltiples monedas)"
    assert _find_summary_value(workbook, "Monto neto:") == "No aplica (múltiples monedas)"

def test_excel_export_escapes_formula_prefixed_text_cells():
    transactions = [
        {
            "date": "2024-01-01",
            "description": "=1+1",
            "amount": 10.0,
            "balance": 10.0,
            "transaction_type": "+Credit",
            "bank": "-bbva",
            "account": "@1",
            "currency": "\tARS",
            "scope_label": "\rscope",
            "product_type": "\nproduct",
            "linked_account": "=linked",
            "source_file": "+file.pdf",
        }
    ]
    summary = _build_summary(
        total_transactions=1,
        selected_scopes=[{"label": "=scope", "product_type": "@product", "currency": "+ARS"}],
    )
    summary["errors"] = ["-bad"]

    workbook_bytes = ExcelGenerator().create_excel_file(transactions, summary)
    workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
    worksheet = workbook["Movimientos"]
    headers = [cell.value for cell in worksheet[1]]
    expected_values = {
        "Descripción": "'=1+1",
        "Tipo": "'+Credit",
        "Banco": "'-bbva",
        "Cuenta": "'@1",
        "Moneda": "'\tARS",
        "Entidad": "'\nscope",
        "Tipo de producto": "'\nproduct",
        "Cuenta vinculada": "'=linked",
        "Archivo de origen": "'+file.pdf",
    }

    for header, expected_value in expected_values.items():
        cell = worksheet.cell(row=2, column=headers.index(header) + 1)
        assert cell.data_type != "f"
        assert cell.value == expected_value

    summary_values = [cell.value for row in workbook["Resumen"].iter_rows() for cell in row]
    assert "'=scope" in summary_values
    assert "'@product" in summary_values
    assert "'+ARS" in summary_values
    assert "'-bad" in summary_values


def test_csv_export_escapes_formula_prefixed_text_cells():
    transactions = [
        {
            "date": "2024-01-01",
            "description": "=1+1",
            "amount": 10.0,
            "balance": 10.0,
            "transaction_type": "+Credit",
            "bank": "-bbva",
            "account": "@1",
            "currency": "\tARS",
            "scope_label": "\rscope",
            "product_type": "\nproduct",
            "linked_account": "=linked",
            "source_file": "+file.pdf",
        }
    ]

    csv_data = _build_csv_export(transactions)
    rows = list(csv.DictReader(io.StringIO(csv_data, newline="")))

    assert len(rows) == 1
    row = rows[0]
    assert row["Fecha"] == "2024-01-01"
    assert row["Monto"] == "10.0"
    assert row["Saldo"] == "10.0"
    assert row["Descripción"] == "'=1+1"
    assert row["Tipo"] == "'+Credit"
    assert row["Banco"] == "'-bbva"
    assert row["Cuenta"] == "'@1"
    assert row["Moneda"] == "'\tARS"
    assert row["Entidad"] == "'\rscope"
    assert row["Tipo de producto"] == "'\nproduct"
    assert row["Cuenta vinculada"] == "'=linked"
    assert row["Archivo de origen"] == "'+file.pdf"


def test_excel_transactions_sheet_formats_amounts_by_header_without_balance():
    transactions = [
        {
            "date": "2024-01-01",
            "description": "Debito",
            "amount": -25.0,
            "transaction_type": "Debit",
            "bank": "bbva",
            "account": "1",
            "currency": "ARS",
        }
    ]

    workbook_bytes = ExcelGenerator().create_excel_file(transactions, _build_summary(total_transactions=1))
    workbook = load_workbook(io.BytesIO(workbook_bytes))
    worksheet = workbook["Movimientos"]
    headers = [cell.value for cell in worksheet[1]]

    assert "Saldo" not in headers
    assert headers[2] == "Monto"
    assert headers[3] == "Tipo"
    assert worksheet["C2"].number_format == "#,##0.00"
    assert worksheet["D2"].value == "Débito"
    assert worksheet["D2"].number_format == "General"


def test_excel_monthly_summaries_keep_currencies_separate():
    transactions = [
        {
            "date": "2024-01-01",
            "description": "Ingreso ARS",
            "amount": 100000.0,
            "balance": 100000.0,
            "transaction_type": "Credit",
            "bank": "brubank",
            "account": "1",
            "currency": "ARS",
        },
        {
            "date": "2024-01-02",
            "description": "Ingreso USD",
            "amount": 100.0,
            "balance": 100.0,
            "transaction_type": "Credit",
            "bank": "brubank",
            "account": "2",
            "currency": "USD",
        },
        {
            "date": "2024-01-03",
            "description": "Debito USD",
            "amount": -10.0,
            "balance": 90.0,
            "transaction_type": "Debit",
            "bank": "brubank",
            "account": "2",
            "currency": "USD",
        },
    ]

    workbook_bytes = ExcelGenerator().create_excel_file(transactions, _build_summary(total_transactions=3))
    workbook = load_workbook(io.BytesIO(workbook_bytes))
    monthly_sheet = workbook["Resumen Mensual"]
    headers = [cell.value for cell in monthly_sheet[2]]
    rows_by_currency = {
        row[headers.index("Moneda")].value: row
        for row in monthly_sheet.iter_rows(min_row=3, max_row=4)
    }

    assert headers == [
        "Mes",
        "Moneda",
        "Total de créditos",
        "Total de débitos",
        "Monto neto",
        "Cantidad de movimientos",
        "Promedio por movimiento",
    ]
    assert rows_by_currency["ARS"][headers.index("Total de créditos")].value == 100000.0
    assert rows_by_currency["ARS"][headers.index("Monto neto")].value == 100000.0
    assert rows_by_currency["USD"][headers.index("Total de créditos")].value == 100.0
    assert rows_by_currency["USD"][headers.index("Total de débitos")].value == 10.0
    assert rows_by_currency["USD"][headers.index("Monto neto")].value == 90.0

    analysis_sheet = workbook["Análisis"]
    assert [cell.value for cell in analysis_sheet[4][6:11]] == ["Mes", "Moneda", "Ingresos", "Egresos", "Neto"]


def test_excel_monthly_summaries_format_single_currency_month_as_text():
    transactions = [
        {
            "date": "2026-06-01",
            "description": "Ingreso",
            "amount": 12388.823,
            "balance": 12388.823,
            "transaction_type": "Credit",
            "bank": "brubank",
            "account": "1",
            "currency": "ARS",
        },
        {
            "date": "2026-06-02",
            "description": "Debito",
            "amount": -6237.815,
            "balance": 6151.008,
            "transaction_type": "Debit",
            "bank": "brubank",
            "account": "1",
            "currency": "ARS",
        },
    ]

    workbook_bytes = ExcelGenerator().create_excel_file(transactions, _build_summary(total_transactions=2))
    workbook = load_workbook(io.BytesIO(workbook_bytes))

    assert workbook["Análisis"]["G5"].value == "2026-06"
    assert workbook["Resumen Mensual"]["A3"].value == "2026-06"


def test_project_metadata_uses_product_identity():
    with Path("pyproject.toml").open("rb") as handle:
        pyproject = tomllib.load(handle)

    project = pyproject["project"]
    assert project["name"] == "bank-statement-parser"
    assert "bank statement" in project["description"].lower()
    assert "Add your description" not in project["description"]
