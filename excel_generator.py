import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import io
import re
from numbers import Real
from typing import List, Dict, Any
from datetime import datetime

from reconciliation_excel import append_reconciliation_summary, create_reconciliation_sheet

from utils import (
    escape_spreadsheet_formula_value,
    format_currency,
    get_transaction_currencies,
    resolve_single_currency,
    user_facing_column_label,
    user_facing_column_order,
    user_facing_product_type,
    user_facing_transaction_value,
)



class ExcelGenerator:
    """Generates structured Excel files from transaction data."""
    
    def __init__(self):
        self.workbook = None
        self.transactions_df = None
        self.reconciliations: List[Dict[str, Any]] = []
    
    def create_excel_file(
        self,
        transactions: List[Dict[str, Any]],
        summary: Dict[str, Any],
        reconciliations: List[Dict[str, Any]] | None = None,
    ) -> bytes:
        """
        Create a comprehensive Excel file with transaction data and analysis.
        
        Args:
            transactions: List of transaction dictionaries
            summary: Processing summary information
            
        Returns:
            Excel file as bytes
        """
        # Create DataFrame
        self.transactions_df = pd.DataFrame(transactions)
        self.reconciliations = [dict(record) for record in (reconciliations or [])]
        
        # Create workbook
        self.workbook = Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        # Create sheets
        self._create_summary_sheet(summary)
        create_reconciliation_sheet(self.workbook, self.reconciliations, self._safe_excel_value)
        self._create_transactions_sheet()
        self._create_scope_transaction_sheets()
        self._create_analysis_sheet()
        self._create_monthly_summary_sheet()
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        self.workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()

    def _transaction_currency_list(self) -> List[str]:
        if self.transactions_df is None or self.transactions_df.empty:
            return []
        return get_transaction_currencies(self.transactions_df.to_dict("records"))

    def _single_transaction_currency(self) -> str | None:
        if self.transactions_df is None or self.transactions_df.empty:
            return None
        return resolve_single_currency(self.transactions_df.to_dict("records"))

    def _display_financial_amount(self, amount: float, multiple_currency_label: str) -> str:
        currencies = self._transaction_currency_list()
        if len(currencies) > 1:
            return multiple_currency_label
        return format_currency(amount, self._single_transaction_currency())

    def _safe_excel_value(self, value: Any) -> Any:
        return escape_spreadsheet_formula_value(value)

    def _set_cell_value(self, worksheet, coordinate: str, value: Any):
        worksheet[coordinate] = self._safe_excel_value(value)

    def _append_dataframe_rows(self, worksheet, dataframe: pd.DataFrame):
        for row in dataframe_to_rows(dataframe, index=False, header=True):
            worksheet.append([self._safe_excel_value(value) for value in row])

    def _prepare_transaction_export_dataframe(self, dataframe: pd.DataFrame) -> pd.DataFrame:
        df_export = dataframe.copy()
        if 'date' in df_export.columns:
            df_export['date'] = pd.to_datetime(df_export['date'], errors='coerce')

        existing_columns = user_facing_column_order(df_export.columns)
        df_export = df_export[existing_columns]
        for column in df_export.columns:
            df_export[column] = df_export[column].map(
                lambda value, column=column: user_facing_transaction_value(column, value)
            )
        return df_export.rename(columns={column: user_facing_column_label(column) for column in df_export.columns})

    def _format_transaction_worksheet(self, worksheet):
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        header_columns = {cell.value: cell.column for cell in worksheet[1]}
        amount_columns = {
            column
            for column in [
                header_columns.get(user_facing_column_label('amount')),
                header_columns.get(user_facing_column_label('balance')),
            ]
            if column is not None
        }
        date_column = header_columns.get(user_facing_column_label('date'))

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border

                if cell.column in amount_columns:
                    cell.number_format = '#,##0.00'
                    if isinstance(cell.value, Real) and not isinstance(cell.value, bool) and cell.value < 0:
                        cell.font = Font(color="FF0000")

                if cell.column == date_column and cell.value:
                    cell.number_format = 'DD/MM/YYYY'

        self._auto_adjust_columns(worksheet, max_width=50)
        worksheet.freeze_panes = 'A2'

    def _auto_adjust_columns(self, worksheet, max_width: int):
        for column in worksheet.columns:
            first = next((c for c in column if not isinstance(c, MergedCell)), column[0])
            col_letter = get_column_letter(first.column)
            lengths = [len(str(c.value)) for c in column if c.value]
            max_length = max(lengths) if lengths else 0
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, max_width)

    def _currency_bucket_series(self, dataframe: pd.DataFrame) -> pd.Series:
        if 'currency' not in dataframe.columns:
            return pd.Series(["N/A"] * len(dataframe), index=dataframe.index)
        return (
            dataframe['currency']
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
            .replace("", "N/A")
        )

    def _monthly_summary_rows(self, dataframe: pd.DataFrame) -> tuple[List[Dict[str, Any]], bool]:
        df_copy = dataframe.copy()
        df_copy['date'] = pd.to_datetime(df_copy['date'], errors='coerce')
        df_copy = df_copy.dropna(subset=['date'])

        if df_copy.empty:
            return [], False

        df_copy['year_month'] = df_copy['date'].dt.to_period('M')
        df_copy['currency_bucket'] = self._currency_bucket_series(df_copy)
        split_by_currency = df_copy['currency_bucket'].nunique(dropna=False) > 1
        group_columns = ['year_month', 'currency_bucket'] if split_by_currency else 'year_month'

        monthly_data = []
        for group_key, group in df_copy.groupby(group_columns, sort=True):
            if split_by_currency:
                month, currency = group_key
            else:
                if isinstance(group_key, tuple) and len(group_key) == 1:
                    group_key = group_key[0]
                month = group_key
                currency = None

            credits = group[group['amount'] > 0]['amount'].sum()
            debits = abs(group[group['amount'] < 0]['amount'].sum())
            net = credits - debits
            row = {
                'Month': str(month),
                'Total Credits': credits,
                'Total Debits': debits,
                'Net Amount': net,
                'Transaction Count': len(group),
                'Average Transaction': group['amount'].mean(),
            }
            if split_by_currency:
                row = {
                    'Month': str(month),
                    'Currency': currency,
                    **{key: value for key, value in row.items() if key != 'Month'},
                }
            monthly_data.append(row)

        return monthly_data, split_by_currency
    
    def _create_summary_sheet(self, summary: Dict[str, Any]):
        """Create summary sheet with processing information."""
        ws = self.workbook.create_sheet("Resumen", 0)
        
        # Title
        ws['A1'] = "Resumen de Procesamiento de Extractos Bancarios"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')
        
        # Processing date
        ws['A3'] = "Fecha de procesamiento:"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # File statistics
        ws['A5'] = "Archivos"
        ws['A5'].font = Font(bold=True, size=12)
        
        ws['A6'] = "Total de archivos cargados:"
        ws['B6'] = summary['total_files']
        
        ws['A7'] = "Procesados correctamente:"
        ws['B7'] = summary['successful_files']
        
        ws['A8'] = "Archivos con error:"
        ws['B8'] = summary['failed_files']
        
        # Transaction statistics
        ws['A10'] = "Movimientos"
        ws['A10'].font = Font(bold=True, size=12)
        
        ws['A11'] = "Total de movimientos:"
        ws['B11'] = summary['total_transactions']
        
        ws['A12'] = "Bancos detectados:"
        ws['B12'] = len(summary['banks_detected'])
        
        ws['A13'] = "Nombres de bancos:"
        self._set_cell_value(ws, 'B13', ", ".join(sorted(summary['banks_detected'])))

        currencies = self._transaction_currency_list()
        ws['A14'] = "Monedas detectadas:"
        self._set_cell_value(ws, 'B14', ", ".join(currencies) if currencies else "N/A")

        selected_scopes = summary.get('selected_scopes', [])
        if selected_scopes:
            ws['A16'] = "Entidades seleccionadas"
            ws['A16'].font = Font(bold=True, size=12)
            ws['A17'] = "Entidad"
            ws['B17'] = "Tipo"
            ws['C17'] = "Moneda"
            for cell in [ws['A17'], ws['B17'], ws['C17']]:
                cell.font = Font(bold=True)

            row = 18
            for scope in selected_scopes:
                self._set_cell_value(ws, f'A{row}', scope.get('label', ''))
                self._set_cell_value(ws, f'B{row}', user_facing_product_type(scope.get('product_type', '')))
                self._set_cell_value(ws, f'C{row}', scope.get('currency', ''))
                row += 1
        else:
            row = 16
        content_end_row = row
        
        # Transaction totals
        if not self.transactions_df.empty:
            total_credits = self.transactions_df[self.transactions_df['amount'] > 0]['amount'].sum()
            total_debits = abs(self.transactions_df[self.transactions_df['amount'] < 0]['amount'].sum())
            net_amount = total_credits - total_debits
            
            summary_row = row if selected_scopes else 16
            ws[f'A{summary_row}'] = "Resumen financiero"
            ws[f'A{summary_row}'].font = Font(bold=True, size=12)
            
            ws[f'A{summary_row + 1}'] = "Total de créditos:"
            self._set_cell_value(
                ws,
                f'B{summary_row + 1}',
                self._display_financial_amount(total_credits, "No aplica (múltiples monedas)"),
            )
            
            ws[f'A{summary_row + 2}'] = "Total de débitos:"
            self._set_cell_value(
                ws,
                f'B{summary_row + 2}',
                self._display_financial_amount(total_debits, "No aplica (múltiples monedas)"),
            )
            
            ws[f'A{summary_row + 3}'] = "Monto neto:"
            net_display = self._display_financial_amount(net_amount, "No aplica (múltiples monedas)")
            self._set_cell_value(ws, f'B{summary_row + 3}', net_display)
            if net_display == "No aplica (múltiples monedas)":
                pass
            elif net_amount < 0:
                ws[f'B{summary_row + 3}'].font = Font(color="FF0000")
            else:
                ws[f'B{summary_row + 3}'].font = Font(color="008000")

            content_end_row = summary_row + 3

        reconciliation_row = content_end_row + 2
        content_end_row = append_reconciliation_summary(
            ws,
            self.reconciliations,
            reconciliation_row,
            self._safe_excel_value,
        )
        
        # Errors section
        if summary['errors']:
            error_row = max(20, content_end_row + 2)
            ws[f'A{error_row}'] = "Errores de procesamiento"
            ws[f'A{error_row}'].font = Font(bold=True, size=12, color="FF0000")
            
            row = error_row + 1
            for error in summary['errors']:
                self._set_cell_value(ws, f'A{row}', error)
                ws[f'A{row}'].font = Font(color="FF0000")
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            first = next((c for c in column if not isinstance(c, MergedCell)), column[0])
            col_letter = get_column_letter(first.column)
            lengths = [len(str(c.value)) for c in column if c.value]
            max_length = max(lengths) if lengths else 0
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    def _create_transactions_sheet(self):
        """Create main transactions sheet."""
        ws = self.workbook.create_sheet("Movimientos")
        
        if self.transactions_df.empty:
            ws['A1'] = "No se encontraron movimientos"
            return
        
        df_excel = self._prepare_transaction_export_dataframe(self.transactions_df)
        self._append_dataframe_rows(ws, df_excel)
        self._format_transaction_worksheet(ws)

    def _create_scope_transaction_sheets(self):
        if self.transactions_df is None or self.transactions_df.empty:
            return

        if 'scope_label' not in self.transactions_df.columns:
            return

        grouped = self.transactions_df.dropna(subset=['scope_label']).groupby(['bank', 'scope_label'], sort=True)
        used_titles: set[str] = set(self.workbook.sheetnames)
        for (bank, scope_label), group in grouped:
            title = self._unique_sheet_title(f"{bank[:8]} {scope_label}", used_titles)
            used_titles.add(title)
            ws = self.workbook.create_sheet(title)
            df_scope = self._prepare_transaction_export_dataframe(group)
            self._append_dataframe_rows(ws, df_scope)
            self._format_transaction_worksheet(ws)

    def _unique_sheet_title(self, base_title: str, used_titles: set[str]) -> str:
        cleaned = "".join(" " if character in '[]:*?/\\\\' else character for character in base_title)
        cleaned = re.sub(r'\s+', ' ', cleaned).strip() or "Entidad"
        cleaned = cleaned[:31].rstrip() or "Entidad"
        candidate = cleaned
        index = 2
        normalized_used_titles = {title.lower() for title in used_titles}
        while candidate.lower() in normalized_used_titles:
            suffix = f" {index}"
            candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
            index += 1
        return candidate
    
    def _create_analysis_sheet(self):
        """Create analysis sheet with charts and insights."""
        ws = self.workbook.create_sheet("Análisis")
        
        if self.transactions_df.empty:
            ws['A1'] = "No hay datos disponibles para analizar"
            return
        
        # Transaction type analysis
        ws['A1'] = "Análisis de Movimientos"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Count by transaction type
        type_counts = self.transactions_df['transaction_type'].value_counts()
        
        ws['A3'] = "Resumen por tipo de movimiento"
        ws['A3'].font = Font(bold=True)
        
        row = 4
        for trans_type, count in type_counts.items():
            self._set_cell_value(ws, f'A{row}', user_facing_transaction_value("transaction_type", trans_type))
            ws[f'B{row}'] = count
            row += 1
        
        # Bank analysis
        if 'bank' in self.transactions_df.columns:
            bank_counts = self.transactions_df['bank'].value_counts()
            
            ws['D3'] = "Movimientos por banco"
            ws['D3'].font = Font(bold=True)
            
            row = 4
            for bank, count in bank_counts.items():
                self._set_cell_value(ws, f'D{row}', bank)
                ws[f'E{row}'] = count
                row += 1
        
        # Monthly spending analysis
        if 'date' in self.transactions_df.columns:
            monthly_data, split_by_currency = self._monthly_summary_rows(self.transactions_df)
            
            ws['G3'] = "Resumen mensual"
            ws['G3'].font = Font(bold=True)

            headers = ["Mes"]
            if split_by_currency:
                headers.append("Moneda")
            headers.extend(["Ingresos", "Egresos", "Neto"])
            start_column = 7
            for offset, header in enumerate(headers):
                cell = ws.cell(row=4, column=start_column + offset, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            row = 5
            for monthly_row in monthly_data:
                values = [monthly_row["Month"]]
                if split_by_currency:
                    values.append(monthly_row["Currency"])
                values.extend([
                    monthly_row["Total Credits"],
                    monthly_row["Total Debits"],
                    monthly_row["Net Amount"],
                ])
                for offset, value in enumerate(values):
                    cell = ws.cell(row=row, column=start_column + offset, value=self._safe_excel_value(value))
                    if headers[offset] in {"Ingresos", "Egresos", "Neto"}:
                        cell.number_format = '#,##0.00'
                    if headers[offset] == "Neto" and isinstance(value, Real) and not isinstance(value, bool) and value < 0:
                        cell.font = Font(color="FF0000")
                
                row += 1
    
    def _create_monthly_summary_sheet(self):
        """Create monthly summary sheet."""
        ws = self.workbook.create_sheet("Resumen Mensual")
        
        if self.transactions_df.empty:
            ws['A1'] = "No hay datos disponibles para el resumen mensual"
            return
        
        # Process data
        monthly_data, split_by_currency = self._monthly_summary_rows(self.transactions_df)

        if not monthly_data:
            ws['A1'] = "No se encontraron fechas válidas para el resumen mensual"
            return
        
        # Create DataFrame and add to sheet
        monthly_df = pd.DataFrame(monthly_data).rename(columns={
            "Month": "Mes",
            "Currency": "Moneda",
            "Total Credits": "Total de créditos",
            "Total Debits": "Total de débitos",
            "Net Amount": "Monto neto",
            "Transaction Count": "Cantidad de movimientos",
            "Average Transaction": "Promedio por movimiento",
        })
        
        # Add title
        ws['A1'] = "Resumen Financiero Mensual"
        ws['A1'].font = Font(size=14, bold=True)
        title_end_column = 'G' if split_by_currency else 'F'
        ws.merge_cells(f'A1:{title_end_column}1')
        
        # Add data
        self._append_dataframe_rows(ws, monthly_df)
        
        # Format header (row 2 since we only added a title row)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Format data rows
        header_columns = {cell.value: cell.column for cell in ws[2]}
        monetary_columns = {
            column
            for column in [
                header_columns.get('Total de créditos'),
                header_columns.get('Total de débitos'),
                header_columns.get('Monto neto'),
                header_columns.get('Promedio por movimiento'),
            ]
            if column is not None
        }
        count_column = header_columns.get('Cantidad de movimientos')

        for row in ws.iter_rows(min_row=3):
            for cell in row:
                # Format monetary columns
                if cell.column in monetary_columns:
                    cell.number_format = '#,##0.00'
                    if isinstance(cell.value, Real) and not isinstance(cell.value, bool) and cell.value < 0:
                        cell.font = Font(color="FF0000")
                
                # Format count column
                elif cell.column == count_column:
                    cell.number_format = '#,##0'
        
        # Auto-adjust column widths
        for column in ws.columns:
            first = next((c for c in column if not isinstance(c, MergedCell)), column[0])
            col_letter = get_column_letter(first.column)
            lengths = [len(str(c.value)) for c in column if c.value]
            max_length = max(lengths) if lengths else 0
            ws.column_dimensions[col_letter].width = min(max_length + 2, 20)
